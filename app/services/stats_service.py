# -*- coding: utf-8 -*-
"""扫码统计服务。"""

import io
from datetime import datetime, timedelta, timezone

from sqlalchemy import Date, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.live_code import LiveCode
from app.models.scan_log import ScanLog


class StatsService:
    """扫码统计与导出。"""

    def __init__(self, session: AsyncSession):
        self.session = session

    async def get_system_overview(self) -> dict[str, int | float]:
        """获取首页实时概览数据。"""
        live_code_count_stmt = select(func.count(LiveCode.id)).where(
            LiveCode.is_deleted == False
        )
        active_live_code_count_stmt = select(func.count(LiveCode.id)).where(
            LiveCode.is_deleted == False,
            LiveCode.is_active == True,
        )
        scan_count_stmt = (
            select(func.count(ScanLog.id))
            .join(LiveCode, ScanLog.live_code_id == LiveCode.id)
            .where(LiveCode.is_deleted == False)
        )

        live_code_count = (await self.session.execute(live_code_count_stmt)).scalar() or 0
        active_live_code_count = (
            (await self.session.execute(active_live_code_count_stmt)).scalar() or 0
        )
        total_scan_count = (await self.session.execute(scan_count_stmt)).scalar() or 0
        availability_rate = (
            round(active_live_code_count / live_code_count * 100, 1)
            if live_code_count else 0.0
        )
        return {
            "live_code_count": live_code_count,
            "total_scan_count": total_scan_count,
            "availability_rate": availability_rate,
        }

    async def record_scan(
        self,
        live_code_id: int,
        ip_address: str | None = None,
        user_agent: str | None = None,
        referer: str | None = None,
        extra_data: dict | None = None,
    ) -> ScanLog:
        """记录一次扫码到 ScanLog。"""
        log = ScanLog(
            live_code_id=live_code_id,
            ip_address=ip_address,
            user_agent=user_agent,
            referer=referer,
            extra_data=extra_data or {},
        )
        self.session.add(log)
        await self.session.flush()
        return log

    async def get_total(self, live_code_id: int) -> int:
        """获取扫码总量（从 ScanLog 表统计，保证准确性）。"""
        stmt = select(func.count(ScanLog.id)).where(
            ScanLog.live_code_id == live_code_id
        )
        result = await self.session.execute(stmt)
        return result.scalar() or 0

    async def get_today(self, live_code_id: int) -> int:
        """今日扫码数。"""
        today_start = datetime.now(timezone.utc).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        stmt = select(func.count(ScanLog.id)).where(
            ScanLog.live_code_id == live_code_id,
            ScanLog.scanned_at >= today_start,
        )
        result = await self.session.execute(stmt)
        return result.scalar() or 0

    async def get_yesterday(self, live_code_id: int) -> int:
        """昨日扫码数。"""
        now = datetime.now(timezone.utc)
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        yesterday_start = today_start - timedelta(days=1)
        stmt = select(func.count(ScanLog.id)).where(
            ScanLog.live_code_id == live_code_id,
            ScanLog.scanned_at >= yesterday_start,
            ScanLog.scanned_at < today_start,
        )
        result = await self.session.execute(stmt)
        return result.scalar() or 0

    async def get_daily_trend(
        self, live_code_id: int, days: int = 30
    ) -> list[dict]:
        """获取按天趋势（最近 N 天）。"""
        since = datetime.now(timezone.utc) - timedelta(days=days)
        stmt = (
            select(
                func.date(ScanLog.scanned_at).label("d"),
                func.count(ScanLog.id).label("cnt"),
            )
            .where(
                ScanLog.live_code_id == live_code_id,
                ScanLog.scanned_at >= since,
            )
            .group_by(func.date(ScanLog.scanned_at))
            .order_by(func.date(ScanLog.scanned_at))
        )
        result = await self.session.execute(stmt)
        rows = result.all()

        # 填充缺失日期（无扫码的日期返回 0）
        trend_map = {str(row[0]): row[1] for row in rows}
        trend = []
        for i in range(days - 1, -1, -1):
            d = (datetime.now(timezone.utc) - timedelta(days=i)).strftime("%Y-%m-%d")
            trend.append({"date": d, "count": trend_map.get(d, 0)})
        return trend

    async def export_excel(
        self, live_code_id: int, days: int = 30
    ) -> bytes:
        """导出为 Excel（openpyxl）。"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        trend = await self.get_daily_trend(live_code_id, days=days)

        wb = Workbook()
        ws = wb.active
        ws.title = "扫码统计"

        # 表头
        headers = ["日期", "扫码次数"]
        header_font = Font(bold=True, size=12)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        for row_idx, item in enumerate(trend, 2):
            ws.cell(row=row_idx, column=1, value=item["date"])
            ws.cell(row=row_idx, column=2, value=item["count"])

        # 自适应列宽
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 12

        # 写入 bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()